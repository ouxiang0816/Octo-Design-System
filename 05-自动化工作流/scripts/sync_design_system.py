#!/usr/bin/env python3
from __future__ import annotations

import argparse
import copy
import html
import json
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - environment guard
    raise SystemExit("Missing dependency: openpyxl. Install it for Python 3 before running ds sync.") from exc


ROOT = Path(__file__).resolve().parents[2]
EXCEL_DIR = ROOT / "01-excel表"
DOCS_DIR = ROOT / "02-markdown文档"
COMPONENTS_DIR = ROOT / "03-开发组件"
PREVIEW_DIR = ROOT / "04-预览网站"
DOC_COMPONENTS_DIR = DOCS_DIR / "components"
WORKFLOW_GENERATED_DIR = ROOT / "05-自动化工作流" / "generated"

EXCEL_SOURCE = EXCEL_DIR / "设计系统组件清单.synced.xlsx"
EXCEL_SYNCED = EXCEL_DIR / "设计系统组件清单.synced.xlsx"
EXCEL_BACKUP = EXCEL_DIR / "设计系统组件清单.full-backup.xlsx"
# Numbers source file — auto-synced to xlsx before any processing
NUMBERS_SOURCE_PATH = EXCEL_DIR / "设计系统组件清单V1.synced.numbers"
BASE_MD = DOCS_DIR / "BASE.md"
COMPONENTS_MD = DOCS_DIR / "COMPONENTS.md"
REGISTRY_TS = PREVIEW_DIR / "src/generated/design-system-registry.ts"
REGISTRY_JSON = WORKFLOW_GENERATED_DIR / "design-system-registry.json"
COMPONENT_DOC_INDEX = DOC_COMPONENTS_DIR / "README.md"
HTML_OUT = PREVIEW_DIR / "Design System Visual Guide.html"
FIGMA_CACHE_DIR = DOCS_DIR / "figma"
AI_REFERENCE_MD = DOCS_DIR / "AI_REFERENCE.md"

SYNC_INDEX_START = "<!-- DS_SYNC_INDEX_START -->"
SYNC_INDEX_END = "<!-- DS_SYNC_INDEX_END -->"
AUTO_BLOCK_START = "<!-- DS_AUTO_START {id} -->"
AUTO_BLOCK_END = "<!-- DS_AUTO_END {id} -->"
SPEC_BLOCK_START = "<!-- DS_SPEC_START -->"
SPEC_BLOCK_END = "<!-- DS_SPEC_END -->"

FIGMA_FETCH_STATUSES = frozenset({"需要", "新增", "修改"})

CATEGORY_SHEETS = {
    "🔵 基础类": ("basic", "基础类"),
    "📦 容器类": ("container", "容器类"),
    "🧭 导航类": ("navigation", "导航类"),
    "📝 表单类": ("form", "表单类"),
    "📊 数据展示类": ("data-display", "数据展示类"),
    "🔔 反馈类": ("feedback", "反馈类"),
}

CATEGORY_LABELS = {
    "foundation": "Foundation",
    "basic": "基础类",
    "container": "容器类",
    "navigation": "导航类",
    "form": "表单类",
    "data-display": "数据展示类",
    "feedback": "反馈类",
}

COMPONENTS_SOURCE = str((COMPONENTS_DIR / "components.tsx").relative_to(ROOT))
DEMOS_SOURCE = str((COMPONENTS_DIR / "interactiveDemos.tsx").relative_to(ROOT))

IMPLEMENTED_COMPONENTS = {
    "component.button": ("Button", COMPONENTS_SOURCE, DEMOS_SOURCE),
    "component.icon-button": ("IconButton", COMPONENTS_SOURCE, DEMOS_SOURCE),
    "component.text-link": ("TextLink", COMPONENTS_SOURCE, DEMOS_SOURCE),
    "component.input": ("Input", COMPONENTS_SOURCE, DEMOS_SOURCE),
    "component.search": ("SearchInput", COMPONENTS_SOURCE, DEMOS_SOURCE),
    "component.select": ("SelectBox", COMPONENTS_SOURCE, DEMOS_SOURCE),
    "component.checkbox": ("Checkbox", COMPONENTS_SOURCE, DEMOS_SOURCE),
    "component.radio": ("Radio", COMPONENTS_SOURCE, DEMOS_SOURCE),
    "component.switch": ("Switch", COMPONENTS_SOURCE, DEMOS_SOURCE),
    "component.tag-chip": ("Tag", COMPONENTS_SOURCE, DEMOS_SOURCE),
    "component.top-navigation": ("TopNavigation", COMPONENTS_SOURCE, DEMOS_SOURCE),
    "component.side-navigation": ("SideNavigation", COMPONENTS_SOURCE, DEMOS_SOURCE),
    "component.breadcrumb": ("Breadcrumb", COMPONENTS_SOURCE, DEMOS_SOURCE),
    "component.tabs": ("Tabs", COMPONENTS_SOURCE, DEMOS_SOURCE),
    "component.steps": ("Steps", COMPONENTS_SOURCE, DEMOS_SOURCE),
    "component.badge": ("Badge", COMPONENTS_SOURCE, DEMOS_SOURCE),
    "component.avatar": ("Avatar", COMPONENTS_SOURCE, DEMOS_SOURCE),
    "component.drawer-sheet": ("DrawerPreview", COMPONENTS_SOURCE, DEMOS_SOURCE),
    "component.步进器": ("Stepper", COMPONENTS_SOURCE, DEMOS_SOURCE),
    "component.file-upload": ("FileUpload", COMPONENTS_SOURCE, DEMOS_SOURCE),
    "component.file-card": ("FileCard", COMPONENTS_SOURCE, DEMOS_SOURCE),
    "component.toast-message": ("Toast", COMPONENTS_SOURCE, DEMOS_SOURCE),
}


@dataclass
class MdSection:
    file: str
    title: str
    anchor: str
    text: str
    line: int


@dataclass
class Entry:
    id: str
    category: str
    name: str
    cn_name: str
    status: str
    figma_url: str | None = None
    sheet_name: str | None = None
    sheet_row: int | None = None
    md: MdSection | None = None
    component: dict[str, str] | None = None
    preview: dict[str, Any] | None = None
    sync_status: str = "missing-md"
    sync_error: str = ""
    management: dict[str, str] = field(default_factory=dict)
    figma: dict[str, Any] | None = None


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def relative_path(path: Path) -> str:
    return str(path.relative_to(ROOT))


def slugify(value: str) -> str:
    value = value.lower().strip()
    value = value.replace("&", " and ")
    value = re.sub(r"[/_]+", " ", value)
    value = re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "-", value)
    value = re.sub(r"-+", "-", value)
    return value.strip("-")


def english_name(title: str) -> str:
    title = re.sub(r"`([^`]+)`", r"\1", title)
    lines = [line.strip() for line in title.splitlines() if line.strip()]
    title = next((line for line in lines if re.search(r"[A-Za-z]", line)), lines[0] if lines else title)
    match = re.match(r"([A-Za-z0-9 /&+-]+)", title.strip())
    if match:
        return match.group(1).strip(" -/")
    return title.strip()


def cn_name(raw: str) -> str:
    lines = [line.strip() for line in raw.splitlines() if line.strip()]
    if lines:
        primary_name = english_name(raw)
        for line in lines:
            if line != primary_name and not re.search(r"[A-Za-z]", line):
                return line
    if len(lines) >= 2 and lines[1] != english_name(raw):
        return lines[1]
    tail = re.sub(r"^[A-Za-z0-9 /&+-]+", "", raw).strip()
    return tail


def heading_key_name(title: str) -> str:
    lines = [line.strip() for line in title.splitlines() if line.strip()]
    joined = " ".join(lines) if lines else title.strip()
    if re.search(r"[A-Za-z]", joined):
        return english_name(joined)
    parts = [part for part in re.split(r"\s+", joined) if part]
    if len(parts) > 1 and len(set(parts)) == 1:
        return parts[0]
    if lines:
        return lines[0]
    return title.strip()


def component_id(name: str) -> str:
    return f"component.{slugify(heading_key_name(name))}"


def foundation_id(title: str) -> str:
    return f"foundation.{slugify(title)}"


def md_anchor(title: str) -> str:
    return slugify(heading_key_name(title))


def component_doc_path(entry_id: str) -> Path:
    return DOC_COMPONENTS_DIR / f"{entry_id}.md"


def extract_component_doc_title(text: str) -> str:
    match = re.search(r"^#\s+(.+?)\s*$", text, flags=re.MULTILINE)
    return match.group(1).strip() if match else ""


def extract_component_spec(text: str) -> tuple[str, int]:
    pattern = re.compile(
        re.escape(SPEC_BLOCK_START) + r"\s*(.*?)\s*" + re.escape(SPEC_BLOCK_END),
        re.DOTALL,
    )
    match = pattern.search(text)
    if match:
        line = text.count("\n", 0, match.start()) + 1
        return match.group(1).strip(), line

    legacy_heading = re.search(r"^##\s+规范正文\s*$", text, flags=re.MULTILINE)
    if legacy_heading:
        spec = text[legacy_heading.end():].strip()
        spec = re.sub(r"\n---\s*$", "", spec).strip()
        line = text.count("\n", 0, legacy_heading.end()) + 1
        return spec, line

    heading = re.search(r"^#\s+.+?$", text, flags=re.MULTILINE)
    if heading:
        line = text.count("\n", 0, heading.end()) + 1
        return text[heading.end():].strip(), line
    return text.strip(), 1


def render_component_source_doc(title: str, entry_id: str, spec: str) -> str:
    return (
        f"# {title}\n\n"
        "> 组件规范主源。与 `../BASE.md` 配合使用。\n"
        f"> 系统ID：`{entry_id}`。本文件人工维护；`../COMPONENTS.md` 由同步脚本自动聚合生成。\n\n"
        f"{SPEC_BLOCK_START}\n\n"
        f"{spec.strip()}\n\n"
        f"{SPEC_BLOCK_END}\n"
    )


def bootstrap_component_doc_sources() -> None:
    DOC_COMPONENTS_DIR.mkdir(parents=True, exist_ok=True)
    legacy_sections = parse_markdown_sections(COMPONENTS_MD, "component") if COMPONENTS_MD.exists() else {}
    for entry_id, section in legacy_sections.items():
        path = component_doc_path(entry_id)
        if path.exists():
            existing = path.read_text(encoding="utf-8")
            if SPEC_BLOCK_START in existing and SPEC_BLOCK_END in existing:
                continue
            title = extract_component_doc_title(existing) or section.title
            spec, _ = extract_component_spec(existing)
            body = spec or section.text
        else:
            title = section.title
            body = section.text
        path.write_text(render_component_source_doc(title, entry_id, body), encoding="utf-8")


def parse_component_doc_sections() -> dict[str, MdSection]:
    sections: dict[str, MdSection] = {}
    DOC_COMPONENTS_DIR.mkdir(parents=True, exist_ok=True)
    for path in sorted(DOC_COMPONENTS_DIR.glob("component.*.md")):
        text = path.read_text(encoding="utf-8")
        title = extract_component_doc_title(text)
        if not title:
            continue
        body, line = extract_component_spec(text)
        sections[path.stem] = MdSection(
            file=relative_path(path),
            title=title,
            anchor=md_anchor(title),
            text=body,
            line=line,
        )
    return sections


def replace_component_spec(text: str, spec: str) -> str:
    pattern = re.compile(
        re.escape(SPEC_BLOCK_START) + r"\s*.*?\s*" + re.escape(SPEC_BLOCK_END),
        re.DOTALL,
    )
    block = f"{SPEC_BLOCK_START}\n\n{spec.strip()}\n\n{SPEC_BLOCK_END}"
    if pattern.search(text):
        return pattern.sub(block, text)
    title = extract_component_doc_title(text) or "未命名组件"
    return render_component_source_doc(title, component_id(title), spec)


def parse_markdown_sections(path: Path, prefix: str) -> dict[str, MdSection]:
    text = path.read_text(encoding="utf-8")
    matches = list(re.finditer(r"^##\s+(.+?)\s*$", text, flags=re.MULTILINE))
    sections: dict[str, MdSection] = {}
    for index, match in enumerate(matches):
        title = match.group(1).strip()
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        line = text.count("\n", 0, match.start()) + 1
        body = text[start:end].strip()
        if title == "同步索引":
            continue
        if prefix == "component":
            key = component_id(title)
        else:
            key = foundation_id(title)
        sections[key] = MdSection(
            file=path.name,
            title=title,
            anchor=md_anchor(title),
            text=body,
            line=line,
        )
    return sections


def extract_figma_ref(url: str | None) -> dict[str, str] | None:
    if not url:
        return None
    file_match = re.search(r"figma\.com/(?:design|file)/([^/?#]+)", url)
    node_match = re.search(r"node-id=([0-9]+[-:][0-9]+)", url)
    if not file_match or not node_match:
        return None
    node_id = node_match.group(1).replace("-", ":")
    return {"fileKey": file_match.group(1), "nodeId": node_id, "url": url}


def figma_cache_for(entry_id: str, figma_url: str | None) -> dict[str, Any] | None:
    ref = extract_figma_ref(figma_url)
    if ref is None:
        return None
    context_path = FIGMA_CACHE_DIR / f"{entry_id}.context.json"
    screenshot_path = FIGMA_CACHE_DIR / f"{entry_id}.screenshot.png"
    ref["contextPath"] = str(context_path.relative_to(ROOT))
    ref["screenshotPath"] = str(screenshot_path.relative_to(ROOT))
    ref["hasContext"] = context_path.exists()
    ref["hasScreenshot"] = screenshot_path.exists()
    if context_path.exists():
        try:
            payload = json.loads(context_path.read_text(encoding="utf-8"))
            ref["summary"] = payload.get("summary", "")
            ref["extracted"] = payload.get("extracted", {})
        except json.JSONDecodeError:
            ref["summary"] = ""
            ref["extracted"] = {}
            ref["cacheError"] = "Invalid JSON cache"
    return ref


def _sync_numbers_to_xlsx() -> None:
    """If Numbers source exists, refresh xlsx data rows (row 3+, cols A-D) from it."""
    if not NUMBERS_SOURCE_PATH.exists():
        return
    try:
        import numbers_parser  # type: ignore
    except ImportError:
        return
    try:
        doc = numbers_parser.Document(str(NUMBERS_SOURCE_PATH))
        num_sheet_map = {s.name: s for s in doc.sheets}
        wb = openpyxl.load_workbook(EXCEL_SOURCE) if EXCEL_SOURCE.exists() else openpyxl.Workbook()
        for sheet_name in CATEGORY_SHEETS:
            if sheet_name not in num_sheet_map:
                continue
            tbl = num_sheet_map[sheet_name].tables[0]
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
            for r in range(2, tbl.num_rows):  # skip header row(0) and subtitle row(1)
                for c in range(min(tbl.num_cols, 4)):
                    try:
                        val = tbl.cell(r, c).value
                        str_val = str(val).strip() if val is not None else None
                        ws.cell(r + 1, c + 1, str_val or None)
                    except Exception:
                        pass
        wb.save(EXCEL_SOURCE)
    except Exception as exc:
        print(f"⚠ Numbers 同步跳过: {exc}")


def parse_excel_entries() -> list[Entry]:
    _sync_numbers_to_xlsx()
    wb = openpyxl.load_workbook(EXCEL_SOURCE, read_only=True, data_only=True)
    entries: list[Entry] = []
    for sheet_name, (category, _) in CATEGORY_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_index < 3:
                continue
            raw_name = clean_cell(row[0] if len(row) > 0 else "")
            if not raw_name or "·" in raw_name or raw_name == "组件名":
                continue
            name = english_name(raw_name)
            if not name:
                continue
            figma_url = clean_cell(row[2] if len(row) > 2 else "") or None
            entry = Entry(
                id=component_id(raw_name),
                category=category,
                name=name,
                cn_name=cn_name(raw_name),
                status=clean_cell(row[1] if len(row) > 1 else ""),
                figma_url=figma_url,
                sheet_name=sheet_name,
                sheet_row=row_index,
                management={
                    "note": clean_cell(row[3] if len(row) > 3 else ""),
                },
                figma=figma_cache_for(component_id(raw_name), figma_url),
            )
            entries.append(entry)
    return entries


def parse_foundation_entries(base_sections: dict[str, MdSection]) -> list[Entry]:
    entries: list[Entry] = []
    for key, section in base_sections.items():
        entries.append(
            Entry(
                id=key,
                category="foundation",
                name=section.title,
                cn_name="",
                status="md",
                md=section,
                preview={"anchor": section.anchor, "group": "Foundation", "states": []},
                sync_status="synced",
            )
        )
    return entries


def extract_tokens(text: str) -> list[str]:
    patterns = [
        r"#[0-9A-Fa-f]{3,8}",
        r"rgba?\([^)]+\)",
        r"\b\d+(?:\.\d+)?px\b",
        r"\bshadow\.[a-zA-Z0-9_-]+\b",
        r"\bradius\.[a-zA-Z0-9_-]+\b",
    ]
    found: list[str] = []
    for pattern in patterns:
        found.extend(re.findall(pattern, text))
    return found


def normalize_token(token: str) -> str:
    return re.sub(r"\s+", "", token).lower()


def conflict_errors(entry: Entry) -> list[str]:
    # The simplified sheet now carries only management metadata.
    # Detailed design decisions live in Markdown/Figma cache, so notes do not
    # participate in spec conflict detection.
    return []


def merge_entries() -> tuple[list[Entry], list[str]]:
    component_sections = parse_component_doc_sections()
    base_sections = parse_markdown_sections(BASE_MD, "foundation")
    entries = parse_foundation_entries(base_sections) + parse_excel_entries()
    errors: list[str] = []
    seen: dict[str, Entry] = {}

    for entry in entries:
        if entry.id in seen:
            errors.append(f"Duplicate id {entry.id} at {entry.sheet_name}:{entry.sheet_row}")
            entry.sync_status = "conflict"
            continue
        seen[entry.id] = entry

        if entry.id.startswith("component."):
            entry.md = component_sections.get(entry.id)
            if entry.id in IMPLEMENTED_COMPONENTS:
                export_name, component_path, example_path = IMPLEMENTED_COMPONENTS[entry.id]
                entry.component = {
                    "exportName": export_name,
                    "path": component_path,
                    "examplePath": example_path,
                }
            if entry.md is not None:
                entry.preview = {
                    "anchor": entry.md.anchor,
                    "group": CATEGORY_LABELS.get(entry.category, entry.category),
                    "states": preview_states(entry.id),
                }

            conflicts = conflict_errors(entry)
            if conflicts:
                entry.sync_status = "conflict"
                entry.sync_error = "; ".join(conflicts)
                errors.append(
                    f"Conflict {entry.id} at {entry.sheet_name}:{entry.sheet_row}, "
                    f"{entry.md.file if entry.md else 'missing-md'}"
                    f"{':' + str(entry.md.line) if entry.md else ''}: {entry.sync_error}"
                )
            elif entry.md is None:
                entry.sync_status = "missing-md"
                entry.sync_error = "No matching component doc in 02-markdown文档/components/"
            elif entry.figma_url and entry.figma and not entry.figma.get("hasContext"):
                entry.sync_status = "missing-figma"
                entry.sync_error = "Figma URL exists but context cache is missing"
            elif entry.component is None:
                entry.sync_status = "missing-component"
                entry.sync_error = "No React component export registered"
            elif entry.preview is None:
                entry.sync_status = "missing-preview"
                entry.sync_error = "No preview anchor generated"
            else:
                entry.sync_status = "synced"
                entry.sync_error = ""

    return entries, errors


def preview_states(entry_id: str) -> list[str]:
    states = {
        "component.button": ["default", "hover", "active", "disabled"],
        "component.icon-button": ["default", "hover", "active", "disabled"],
        "component.input": ["default", "focus", "error", "disabled"],
        "component.search": ["default", "focus", "with-clear"],
        "component.checkbox": ["unchecked", "checked", "indeterminate", "disabled"],
        "component.radio": ["unchecked", "checked", "disabled"],
        "component.switch": ["off", "on", "disabled"],
        "component.tag-chip": ["blue", "orange", "red", "green", "closable"],
        "component.tabs": ["active", "inactive"],
        "component.steps": ["done", "current", "pending", "error"],
    }
    return states.get(entry_id, ["default"])


def retrieval_keywords(entry: Entry) -> list[str]:
    values = [
        entry.id,
        entry.name,
        entry.cn_name,
        entry.status,
        CATEGORY_LABELS.get(entry.category, entry.category),
        entry.md.title if entry.md else "",
        entry.component["exportName"] if entry.component else "",
    ]
    keywords: list[str] = []
    seen: set[str] = set()
    for value in values:
        normalized = clean_cell(value)
        if normalized and normalized not in seen:
            seen.add(normalized)
            keywords.append(normalized)
    return keywords


def recommended_context_paths(entry: Entry) -> list[str]:
    paths = [relative_path(BASE_MD)]
    if entry.id.startswith("component."):
        if entry.md is not None:
            paths.append(relative_path(component_doc_path(entry.id)))
        else:
            paths.append(relative_path(COMPONENTS_MD))
        if entry.figma and entry.figma.get("hasContext"):
            paths.append(entry.figma["contextPath"])
        if entry.component:
            paths.append(entry.component["path"])
            example_path = entry.component.get("examplePath")
            if example_path and example_path not in paths:
                paths.append(example_path)
    return paths


def docs_payload(entry: Entry) -> dict[str, Any]:
    component_path = None
    if entry.id.startswith("component.") and entry.md is not None:
        component_path = relative_path(component_doc_path(entry.id))
    return {
        "basePath": relative_path(BASE_MD),
        "aggregatePath": relative_path(COMPONENTS_MD) if entry.id.startswith("component.") else None,
        "componentPath": component_path,
        "figmaContextPath": None if entry.figma is None else entry.figma.get("contextPath"),
        "figmaScreenshotPath": None if entry.figma is None else entry.figma.get("screenshotPath"),
    }


def resolve_md_file_path(md_file: str) -> str:
    if "/" in md_file:
        return md_file
    return relative_path(DOCS_DIR / md_file)


def entry_to_json(entry: Entry) -> dict[str, Any]:
    return {
        "id": entry.id,
        "category": entry.category,
        "name": entry.name,
        "cnName": entry.cn_name,
        "status": entry.status,
        "figmaUrl": entry.figma_url,
        "sheetName": entry.sheet_name,
        "sheetRow": entry.sheet_row,
        "md": None
        if entry.md is None
        else {
            "file": entry.md.file,
            "path": resolve_md_file_path(entry.md.file),
            "title": entry.md.title,
            "anchor": entry.md.anchor,
            "line": entry.md.line,
            "summary": summarize_md(entry.md.text),
        },
        "docs": docs_payload(entry),
        "component": entry.component,
        "preview": entry.preview,
        "syncStatus": entry.sync_status,
        "syncError": entry.sync_error,
        "management": entry.management,
        "figma": entry.figma,
        "retrieval": {
            "recommendedContextPaths": recommended_context_paths(entry),
            "keywords": retrieval_keywords(entry),
            "states": [] if entry.preview is None else entry.preview.get("states", []),
            "variantSummary": (
                ""
                if entry.figma is None
                or not isinstance(entry.figma.get("extracted"), dict)
                else str(entry.figma["extracted"].get("variants", ""))
            ),
        },
    }


def summarize_md(text: str) -> str:
    lines = [line.strip() for line in text.splitlines()]
    prose = []
    for line in lines:
        if (
            not line
            or line.startswith("|")
            or line.startswith("---")
            or line.startswith("<!--")
            or line.startswith("###")
        ):
            continue
        prose.append(re.sub(r"\s+", " ", line))
        if len(" ".join(prose)) > 170:
            break
    return " ".join(prose)[:220]


def build_registry_payload(entries: list[Entry]) -> tuple[str, list[dict[str, Any]], list[dict[str, str]]]:
    generated_at = datetime.now(timezone.utc).isoformat(timespec="seconds")
    payload = [entry_to_json(entry) for entry in entries]
    categories = [
        {"id": key, "label": label}
        for key, label in CATEGORY_LABELS.items()
        if any(entry.category == key for entry in entries)
    ]
    return generated_at, payload, categories


def write_registry(entries: list[Entry]) -> None:
    generated_at, payload, categories = build_registry_payload(entries)
    content = (
        "// This file is generated by 05-自动化工作流/scripts/sync_design_system.py. Do not edit by hand.\n\n"
        "export type DesignSystemCategory = 'foundation' | 'basic' | 'container' | 'navigation' | 'form' | 'data-display' | 'feedback'\n"
        "export type SyncStatus = 'synced' | 'missing-figma' | 'missing-md' | 'missing-component' | 'missing-preview' | 'conflict'\n\n"
        "export type DesignSystemEntry = {\n"
        "  id: string\n"
        "  category: DesignSystemCategory\n"
        "  name: string\n"
        "  cnName: string\n"
        "  status: string\n"
        "  figmaUrl: string | null\n"
        "  sheetName: string | null\n"
        "  sheetRow: number | null\n"
        "  md: { file: string; path: string; title: string; anchor: string; line: number; summary: string } | null\n"
        "  docs: { basePath: string; aggregatePath: string | null; componentPath: string | null; figmaContextPath: string | null; figmaScreenshotPath: string | null }\n"
        "  component: { exportName: string; path: string; examplePath: string } | null\n"
        "  preview: { anchor: string; group: string; states: string[] } | null\n"
        "  syncStatus: SyncStatus\n"
        "  syncError: string\n"
        "  management: Record<string, string>\n"
        "  figma: { fileKey: string; nodeId: string; url: string; contextPath: string; screenshotPath: string; hasContext: boolean; hasScreenshot: boolean; summary?: string; extracted?: Record<string, unknown>; cacheError?: string } | null\n"
        "  retrieval: { recommendedContextPaths: string[]; keywords: string[]; states: string[]; variantSummary: string }\n"
        "}\n\n"
        f"export const syncGeneratedAt = {json.dumps(generated_at)}\n"
        f"export const designSystemCategories = {json.dumps(categories, ensure_ascii=False, indent=2)} as const\n"
        f"export const designSystemEntries: DesignSystemEntry[] = {json.dumps(payload, ensure_ascii=False, indent=2)}\n"
    )
    REGISTRY_TS.write_text(content, encoding="utf-8")


def write_registry_json(entries: list[Entry]) -> None:
    generated_at, payload, categories = build_registry_payload(entries)
    WORKFLOW_GENERATED_DIR.mkdir(parents=True, exist_ok=True)
    REGISTRY_JSON.write_text(
        json.dumps(
            {
                "generatedAt": generated_at,
                "categories": categories,
                "entries": payload,
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )


def write_component_doc_index(entries: list[Entry]) -> None:
    DOC_COMPONENTS_DIR.mkdir(parents=True, exist_ok=True)
    component_entries = [entry for entry in entries if entry.id.startswith("component.") and entry.md is not None]
    rows = [
        "| 系统ID | 组件 | 文档 | React | 同步状态 |",
        "|---|---|---|---|---|",
    ]
    for entry in component_entries:
        rows.append(
            f"| `{entry.id}` | {entry.md.title} | [{entry.id}.md]({entry.id}.md) | "
            f"{'`' + entry.component['exportName'] + '`' if entry.component else '—'} | `{entry.sync_status}` |"
        )
    COMPONENT_DOC_INDEX.write_text(
        "\n".join(
            [
                "# Component Docs",
                "",
                "> 本目录中的 `component.*.md` 是单组件规范主源，供团队和 AI 按需读取与维护。",
                "> 本 README 为自动生成索引；推荐流程：先查 `05-自动化工作流/generated/design-system-registry.json`，再加载 `02-markdown文档/BASE.md` 和对应 `component.*.md`。",
                "",
            ]
            + rows
        )
        + "\n",
        encoding="utf-8",
    )


def build_component_index_block(entries: list[Entry]) -> str:
    rows = [
        "| 系统ID | 组件 | Markdown | React | HTML预览 | 同步状态 |",
        "|---|---|---|---|---|---|",
    ]
    component_entries = [entry for entry in entries if entry.id.startswith("component.")]
    for entry in component_entries:
        md_link = "—" if entry.md is None else f"[{entry.md.title}](#{entry.md.anchor})"
        react = "—" if entry.component is None else f"`{entry.component['exportName']}`"
        preview = "—" if entry.preview is None else f"`#{entry.preview['anchor']}`"
        rows.append(
            f"| `{entry.id}` | {entry.name} {entry.cn_name} | {md_link} | {react} | {preview} | `{entry.sync_status}` |"
        )
    return (
        f"{SYNC_INDEX_START}\n\n"
        "## 同步索引\n\n"
        "> 自动生成：由 `npm run ds:sync` 更新，用于连接表格清单、Markdown 规范、React 组件和网页预览。\n\n"
        + "\n".join(rows)
        + f"\n\n{SYNC_INDEX_END}\n"
    )


def write_components_aggregate(entries: list[Entry]) -> None:
    sections: list[str] = []
    component_entries = [entry for entry in entries if entry.id.startswith("component.") and entry.md is not None]
    for entry in component_entries:
        sections.append(f"## {entry.md.title}\n\n{entry.md.text.strip()}")

    content_parts = [
        "# Octo Designer 设计系统 — 基础组件规格",
        "",
        "> 与 BASE.md 配合使用。本文件为自动聚合总表；AI 与团队成员应优先通过 `05-自动化工作流/generated/design-system-registry.json` 定位后，再按需读取 `02-markdown文档/components/component.*.md`。",
        "> 本文件参数以 Figma 实测为准（文件：Octo-Designer，UuRaxW6YNJVqnaxq0ihi6S）。",
        "",
        "## 本文件职责",
        "",
        "- 作为组件规范总表，汇总各组件主规范文档和同步索引。",
        "- 便于整体浏览、校对和对外分享，不再作为组件规范唯一主源。",
        "",
        "## 本文件不负责",
        "",
        "- 不作为 AI 的默认首读入口；常规检索优先走 `registry.json + BASE.md + component.*.md`。",
        "- 不承担团队协作流程说明；流程和命令统一写在 `HOWTO.md`。",
    ]
    if sections:
        content_parts.extend(["", "---", "", "\n\n---\n\n".join(sections)])
    content_parts.extend(["", "---", "", build_component_index_block(entries).strip(), ""])
    COMPONENTS_MD.write_text("\n".join(content_parts), encoding="utf-8")


def write_synced_workbook(entries: list[Entry]) -> None:
    wb = openpyxl.load_workbook(EXCEL_SOURCE)
    by_location = {
        (entry.sheet_name, entry.sheet_row): entry
        for entry in entries
        if entry.sheet_name and entry.sheet_row
    }
    headers = ["系统ID", "Figma节点", "Figma缓存", "MD锚点", "React组件", "HTML预览", "同步状态", "同步错误"]
    for sheet_name in CATEGORY_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        start_col = 5
        for offset, header in enumerate(headers):
            cell = ws.cell(row=1, column=start_col + offset, value=header)
            if ws.cell(row=1, column=1).has_style:
                cell._style = copy.copy(ws.cell(row=1, column=1)._style)
        for row_index in range(3, ws.max_row + 1):
            entry = by_location.get((sheet_name, row_index))
            if entry is None:
                continue
            values = [
                entry.id,
                "" if entry.figma is None else entry.figma["nodeId"],
                "" if entry.figma is None else ("已缓存" if entry.figma.get("hasContext") else "缺少缓存"),
                "" if entry.md is None else f"{entry.md.file}#{entry.md.anchor}",
                "" if entry.component is None else entry.component["exportName"],
                "" if entry.preview is None else f"#{entry.preview['anchor']}",
                entry.sync_status,
                entry.sync_error,
            ]
            for offset, value in enumerate(values):
                ws.cell(row=row_index, column=start_col + offset, value=value)
    wb.save(EXCEL_SYNCED)


def normalize_source_workbook() -> None:
    wb = openpyxl.load_workbook(EXCEL_SOURCE)
    if not EXCEL_BACKUP.exists():
        backup_wb = openpyxl.load_workbook(EXCEL_SOURCE)
        backup_wb.save(EXCEL_BACKUP)
    headers = ["组件名称", "状态", "Figma设计链接", "备注说明"]
    for sheet_name in CATEGORY_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = []
        for row_index in range(3, ws.max_row + 1):
            name = clean_cell(ws.cell(row_index, 1).value)
            if not name:
                continue
            rows.append([
                name,
                clean_cell(ws.cell(row_index, 2).value),
                clean_cell(ws.cell(row_index, 3).value),
                clean_cell(ws.cell(row_index, 12).value) or clean_cell(ws.cell(row_index, 4).value),
            ])
        ws.delete_cols(1, ws.max_column)
        for col_index, header in enumerate(headers, start=1):
            ws.cell(1, col_index, header)
        ws.cell(2, 1, f"{CATEGORY_SHEETS[sheet_name][1]} · 仅维护组件名称、状态、Figma设计链接、备注说明")
        for row_index, row_values in enumerate(rows, start=3):
            for col_index, value in enumerate(row_values, start=1):
                ws.cell(row_index, col_index, value)
    wb.save(EXCEL_SOURCE)


def needs_figma_fetch(status: str) -> bool:
    return any(kw in status for kw in FIGMA_FETCH_STATUSES)


def figma_status(entries: list[Entry]) -> int:
    linked = [entry for entry in entries if entry.figma_url]
    missing = [entry for entry in linked if not entry.figma or not entry.figma.get("hasContext") or not entry.figma.get("hasScreenshot")]
    print("Figma cache status")
    print("Linked components:", len(linked))
    print("Complete cache:", len(linked) - len(missing))
    print("Missing cache:", len(missing))
    for entry in missing:
        print(f"  - {entry.id}: {entry.figma_url}")
    return 0


def figma_needed(entries: list[Entry]) -> int:
    """列出状态含「需要/新增/修改」且缺少 Figma context 缓存的组件，供 Claude Figma MCP 逐个抓取。"""
    candidates = [e for e in entries if e.figma_url and needs_figma_fetch(e.status)]
    missing = [e for e in candidates if not e.figma or not e.figma.get("hasContext")]
    cached = len(candidates) - len(missing)

    print(f"Figma fetch needed  (过滤条件: 状态包含 {'、'.join(sorted(FIGMA_FETCH_STATUSES))})")
    print(f"  候选组件: {len(candidates)}  已缓存: {cached}  待抓取: {len(missing)}")

    if not missing:
        print("\n所有候选组件的 Figma context 均已缓存，无需抓取。")
        return 0

    print("\n待抓取列表（将此输出交给 Claude 执行 Figma MCP fetch）:")
    for entry in missing:
        ref = extract_figma_ref(entry.figma_url)
        file_key = ref["fileKey"] if ref else "—"
        node_id  = ref["nodeId"]  if ref else "—"
        out_ctx  = FIGMA_CACHE_DIR / f"{entry.id}.context.json"
        out_img  = FIGMA_CACHE_DIR / f"{entry.id}.screenshot.png"
        print(f"\n  [{entry.id}]")
        print(f"    name     : {entry.name} {entry.cn_name}")
        print(f"    status   : {entry.status}")
        print(f"    figmaUrl : {entry.figma_url}")
        print(f"    fileKey  : {file_key}")
        print(f"    nodeId   : {node_id}")
        print(f"    save_ctx : {out_ctx.relative_to(ROOT)}")
        print(f"    save_img : {out_img.relative_to(ROOT)}")
    return 0


def format_figma_auto_block(entry: Entry) -> str:
    figma = entry.figma or {}
    extracted = figma.get("extracted", {}) if isinstance(figma.get("extracted"), dict) else {}
    rows = [
        ("来源", entry.figma_url or "—"),
        ("Figma 节点", figma.get("nodeId", "—")),
        ("缓存", figma.get("contextPath", "—")),
        ("截图", figma.get("screenshotPath", "—") if figma.get("hasScreenshot") else "—"),
        ("尺寸", extracted.get("size", "待从 Figma 补充")),
        ("颜色", extracted.get("colors", "待从 Figma 补充")),
        ("布局", extracted.get("layout", "待从 Figma 补充")),
        ("变体/状态", extracted.get("variants", "待从 Figma 补充")),
    ]
    table = "\n".join(f"| {label} | {value} |" for label, value in rows)
    summary = figma.get("summary") or "已登记 Figma 来源，等待或已完成 Codex/Figma MCP 采集。"
    return (
        f"{AUTO_BLOCK_START.format(id=entry.id)}\n\n"
        "### Figma 自动采集\n\n"
        f"> {summary}\n\n"
        "| 项目 | 内容 |\n"
        "|---|---|\n"
        f"{table}\n\n"
        f"{AUTO_BLOCK_END.format(id=entry.id)}"
    )


def update_component_docs_from_figma(entries: list[Entry]) -> None:
    for entry in entries:
        if not entry.figma_url:
            continue
        block = format_figma_auto_block(entry)
        path = component_doc_path(entry.id)
        title = f"{entry.name} {entry.cn_name}".strip()
        if path.exists():
            text = path.read_text(encoding="utf-8")
            spec, _ = extract_component_spec(text)
        else:
            text = render_component_source_doc(
                title,
                entry.id,
                f"<!-- DS_MANUAL_START {entry.id} -->\n\n人工补充区：如需补充业务语义、特殊交互或实现注意事项，请写在此区块内。\n\n<!-- DS_MANUAL_END {entry.id} -->",
            )
            spec, _ = extract_component_spec(text)
        block_pattern = re.compile(
            re.escape(AUTO_BLOCK_START.format(id=entry.id))
            + r".*?"
            + re.escape(AUTO_BLOCK_END.format(id=entry.id)),
            re.DOTALL,
        )
        if block_pattern.search(spec):
            spec = block_pattern.sub(block, spec)
        else:
            spec = f"{block}\n\n{spec.strip()}".strip()
        updated = replace_component_spec(text, spec)
        if not path.exists():
            updated = render_component_source_doc(title, entry.id, spec)
        path.write_text(updated, encoding="utf-8")


def write_static_html(entries: list[Entry]) -> None:
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    nav = []
    body = []
    for category, label in CATEGORY_LABELS.items():
        grouped = [entry for entry in entries if entry.category == category]
        if not grouped:
            continue
        nav.append(f"<a href='#{html.escape(category)}'>{html.escape(label)}</a>")
        cards = []
        for entry in grouped:
            anchor = entry.preview["anchor"] if entry.preview else entry.id.replace(".", "-")
            cards.append(
                "<article class='card' id='{anchor}'>"
                "<div class='card-head'><h3>{title}</h3><span class='status {status}'>{status}</span></div>"
                "<p>{summary}</p>"
                "<dl>"
                "<dt>系统ID</dt><dd><code>{id}</code></dd>"
                "<dt>Markdown</dt><dd>{md}</dd>"
                "<dt>React</dt><dd>{react}</dd>"
                "<dt>Excel</dt><dd>{excel}</dd>"
                "</dl>"
                "</article>".format(
                    anchor=html.escape(anchor),
                    title=html.escape(f"{entry.name} {entry.cn_name}".strip()),
                    status=html.escape(entry.sync_status),
                    summary=html.escape(entry.md and summarize_md(entry.md.text) or entry.sync_error or "待补全规范。"),
                    id=html.escape(entry.id),
                    md=html.escape("—" if entry.md is None else f"{entry.md.file}#{entry.md.anchor}"),
                    react=html.escape("—" if entry.component is None else entry.component["exportName"]),
                    excel=html.escape("—" if entry.sheet_name is None else f"{entry.sheet_name} Row {entry.sheet_row}"),
                )
            )
        body.append(
            f"<section id='{html.escape(category)}'><h2>{html.escape(label)}</h2><div class='grid'>{''.join(cards)}</div></section>"
        )
    html_text = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Octo Designer — Design System Visual Guide</title>
  <style>
    :root {{ --brand:#0067D1; --text:#191919; --muted:#777777; --border:#E5E7EB; --selected:#EFF6FF; --ok:#52C41A; --warn:#FAAD14; --err:#FF4D4F; }}
    * {{ box-sizing: border-box; }}
    body {{ margin:0; font-family:"HarmonyOS Sans", Inter, system-ui, -apple-system, sans-serif; color:var(--text); background:#F5F6F8; }}
    .app {{ display:grid; grid-template-columns:255px minmax(0,1fr); min-height:100vh; }}
    aside {{ position:sticky; top:0; height:100vh; background:#fff; border-right:1px solid var(--border); padding:20px 16px; overflow:auto; }}
    .brand {{ font-weight:700; color:#101828; margin-bottom:4px; }}
    .meta {{ color:var(--muted); font-size:12px; margin-bottom:20px; }}
    nav {{ display:flex; flex-direction:column; gap:4px; }}
    nav a {{ color:var(--text); text-decoration:none; height:34px; display:flex; align-items:center; border-radius:4px; padding:0 10px; font-size:12px; }}
    nav a:hover {{ background:#F3F4F6; }}
    main {{ padding:32px 40px 64px; }}
    h1 {{ margin:0 0 6px; font-size:24px; line-height:32px; color:#101828; }}
    .lead {{ margin:0 0 28px; color:var(--muted); font-size:13px; }}
    section {{ margin-bottom:42px; }}
    h2 {{ font-size:18px; line-height:26px; margin:0 0 14px; padding-bottom:10px; border-bottom:1px solid var(--border); }}
    .grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(280px,1fr)); gap:12px; }}
    .card {{ background:#fff; border:1px solid var(--border); border-radius:8px; padding:16px; }}
    .card-head {{ display:flex; align-items:center; justify-content:space-between; gap:12px; margin-bottom:8px; }}
    h3 {{ margin:0; font-size:15px; line-height:22px; color:#101828; }}
    p {{ margin:0 0 12px; color:#364153; font-size:12px; line-height:20px; min-height:40px; }}
    .status {{ border-radius:4px; padding:2px 7px; font-size:11px; white-space:nowrap; background:#F3F4F6; color:var(--muted); }}
    .status.synced {{ background:#F6FFED; color:var(--ok); }}
    .status.missing-md,.status.missing-component,.status.missing-preview {{ background:#FFFBE6; color:#9A6700; }}
    .status.conflict {{ background:#FFF2F0; color:var(--err); }}
    dl {{ display:grid; grid-template-columns:72px minmax(0,1fr); gap:6px 10px; margin:0; font-size:11px; }}
    dt {{ color:var(--muted); }}
    dd {{ margin:0; min-width:0; overflow-wrap:anywhere; }}
    code {{ color:var(--brand); font-family:ui-monospace,SFMono-Regular,Menlo,monospace; }}
  </style>
</head>
<body>
  <div class="app">
    <aside>
      <div class="brand">Octo Designer</div>
      <div class="meta">Generated {html.escape(generated_at)}</div>
      <nav>{''.join(nav)}</nav>
    </aside>
    <main>
      <h1>设计系统同步预览</h1>
      <p class="lead">由 Excel 清单与 Markdown 规范生成。运行 <code>npm run ds:build</code> 后此文件会被 React 预览构建产物覆盖。</p>
      {''.join(body)}
    </main>
  </div>
</body>
</html>
"""
    HTML_OUT.write_text(html_text, encoding="utf-8")


def report(entries: list[Entry], errors: list[str]) -> int:
    counts: dict[str, int] = {}
    for entry in entries:
        counts[entry.sync_status] = counts.get(entry.sync_status, 0) + 1
    print("Design system sync check")
    print("Entries:", len(entries))
    for status in ["synced", "missing-figma", "missing-md", "missing-component", "missing-preview", "conflict"]:
        print(f"  {status}: {counts.get(status, 0)}")
    missing = [entry for entry in entries if entry.sync_status != "synced" and entry.sync_status != "conflict"]
    if missing:
        print("\nPending mappings:")
        for entry in missing[:30]:
            location = "" if entry.sheet_name is None else f" ({entry.sheet_name}:{entry.sheet_row})"
            print(f"  - {entry.id}{location}: {entry.sync_status} - {entry.sync_error}")
        if len(missing) > 30:
            print(f"  ... {len(missing) - 30} more")
    if errors:
        print("\nErrors:")
        for error in errors:
            print(f"  - {error}")
        return 1
    return 0


def strip_figma_auto_blocks(text: str) -> str:
    cleaned = re.sub(
        r"<!-- DS_AUTO_START [^>]+ -->.*?<!-- DS_AUTO_END [^>]+ -->",
        "",
        text,
        flags=re.DOTALL,
    )
    cleaned = re.sub(
        r"<!-- DS_MANUAL_START [^>]+ -->.*?<!-- DS_MANUAL_END [^>]+ -->",
        "",
        cleaned,
        flags=re.DOTALL,
    )
    return re.sub(r"\n{3,}", "\n\n", cleaned).strip()


def extract_base_block(base_source: str, heading: str) -> str:
    match = re.search(rf"^{re.escape(heading)}\s*$", base_source, re.MULTILINE)
    if not match:
        return ""
    start = match.end()
    next_h2 = re.search(r"^##\s+", base_source[start:], re.MULTILINE)
    end = start + next_h2.start() if next_h2 else len(base_source)
    return base_source[start:end].strip()


def generate_ai_reference(entries: list[Entry]) -> None:
    base_source = BASE_MD.read_text(encoding="utf-8")
    timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    implemented = [e for e in entries if e.component and e.id.startswith("component.")]

    lines: list[str] = [
        "# Octo Design System · AI Reference",
        "",
        f"> 自动生成，请勿手动编辑。最后同步：{timestamp}",
        "> 供 Claude Code / Codex / OpenCode 等 AI 编码工具读取，确保生成代码符合 Octo 设计规范。",
        "",
        "---",
        "",
        "## 使用方式",
        "",
        "在项目的 `CLAUDE.md` / `AGENTS.md` 中引用本文件（GitHub 迁移后替换为 raw URL）：",
        "```",
        "参考 Octo 设计规范：<本文件绝对路径或 GitHub raw URL>",
        "```",
        "",
        "---",
        "",
        "## 可用组件",
        "",
        "| 系统ID | 中文名 | 导出名 | 分类 |",
        "|---|---|---|---|",
    ]

    for entry in implemented:
        category_label = CATEGORY_LABELS.get(entry.category, entry.category)
        lines.append(
            f"| `{entry.id}` | {entry.cn_name} | `{entry.component['exportName']}` | {category_label} |"
        )

    export_names = [e.component["exportName"] for e in implemented]
    lines += ["", "## 导入方式", "", "```tsx"]
    lines.append("import {")
    for name in export_names:
        lines.append(f"  {name},")
    lines.append("} from '@octo/design-system'")
    lines += ["```", "", "---", "", "## 组件规范"]

    for entry in implemented:
        lines += ["", f"### {entry.component['exportName']}（{entry.cn_name}）", ""]
        if entry.md and entry.md.text:
            clean_spec = strip_figma_auto_blocks(entry.md.text)
            if clean_spec:
                lines.append(clean_spec)
        else:
            lines.append("_规范文档待补充。_")

    lines += ["", "---", "", "## Design Tokens", ""]
    for heading in ["## 颜色 Token", "## 状态色（语义色）", "## 字号规格", "## 圆角 Token", "## 关键组件尺寸"]:
        block = extract_base_block(base_source, heading)
        if block:
            lines += [heading, "", block, ""]

    ai_rules = extract_base_block(base_source, "## AI 生成规则")
    if ai_rules:
        lines += ["---", "", "## AI 生成规则", "", ai_rules, ""]

    checklist = extract_base_block(base_source, "## 验收清单")
    if checklist:
        lines += ["## 验收清单", "", checklist, ""]

    AI_REFERENCE_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Sync Octo Designer design system sources.")
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--check", action="store_true", help="validate sources without writing files")
    mode.add_argument("--sync", action="store_true", help="generate registry, index, synced workbook and HTML")
    mode.add_argument("--sheet", action="store_true", help="normalize source workbook to the four management columns")
    mode.add_argument("--figma-status", action="store_true", help="print Figma cache status")
    mode.add_argument("--figma-needed", action="store_true", help="list components (status=需要/新增/修改) missing Figma context cache")
    parser.add_argument("--skip-html", action="store_true", help="do not write Design System Visual Guide.html")
    args = parser.parse_args()

    for required in [EXCEL_SOURCE, BASE_MD]:
        if not required.exists():
            raise SystemExit(f"Required source file not found: {required}")

    if args.sheet:
        normalize_source_workbook()
        print(f"Normalized {EXCEL_SOURCE.name} to four management columns.")
        print(f"Backup: {EXCEL_BACKUP.relative_to(ROOT)}")
        return 0

    if args.figma_status:
        return figma_status(parse_excel_entries())

    if args.figma_needed:
        return figma_needed(parse_excel_entries())

    bootstrap_component_doc_sources()

    if args.sync:
        update_component_docs_from_figma(parse_excel_entries())

    entries, errors = merge_entries()
    exit_code = report(entries, errors)
    if args.check:
        return exit_code
    if errors:
        return exit_code

    write_registry(entries)
    write_registry_json(entries)
    write_component_doc_index(entries)
    write_components_aggregate(entries)
    write_synced_workbook(entries)
    generate_ai_reference(entries)
    if not args.skip_html:
        write_static_html(entries)
    print("\nGenerated:")
    print(f"  - {REGISTRY_TS.relative_to(ROOT)}")
    print(f"  - {REGISTRY_JSON.relative_to(ROOT)}")
    print(f"  - {COMPONENT_DOC_INDEX.relative_to(ROOT)}")
    print(f"  - {COMPONENTS_MD.relative_to(ROOT)}")
    print(f"  - {EXCEL_SYNCED.relative_to(ROOT)}")
    print(f"  - {AI_REFERENCE_MD.relative_to(ROOT)}")
    if not args.skip_html:
        print(f"  - {HTML_OUT.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
